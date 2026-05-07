"""
rebuild_invoice_import_sheet.py
One-shot script: rebuilds the 'Invoice Import' sheet in
RazorSync_Invoice_Tracker.xlsx with the new 7-column layout.

New columns: Invoice # | Customer Name | Service Address | Date |
             Item/Service | Unit Price | Memo

Differences from previous layout:
    - Tax column removed.
    - Unit Price column carries a VLOOKUP formula that pulls Price from the
      Service Items library (column C). Auto-fills when the user picks an
      Item/Service.
    - Data validation (Item/Service dropdown, Customer Name dropdown) applied
      to rows 2..ROW_MAX so user-inserted rows inherit dropdowns.
    - README tab refreshed.
    - Service Items library preserved untouched.

Usage:
    python rebuild_invoice_import_sheet.py [path/to/RazorSync_Invoice_Tracker.xlsx]
"""

import os, sys
from pathlib import Path
from os import replace as _os_replace
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# Must match WRITE_ROW_MAX in sync_to_lookup.py — both define the managed range.
ROW_MAX = 2000

SHEET_NAME = "Invoice Import"
LIB_SHEET  = "Service Items"
README_SHEET = "README"

# Canonical customer names written into a named range so the dropdown does not
# need a hardcoded formula string. To add customers, update this list and rerun.
CUSTOMER_NAMES = [
    "American Homes 4 Rent",
    "Main Street Renewal",
]
CUSTOMER_NAMES_RANGE_NAME = "CustomerNames"
# Column in README sheet used to hold the customer name list (far right, out of way).
CUSTOMER_NAMES_COL = "D"

HEADERS = [
    "Invoice #",
    "Customer Name",
    "Service Address",
    "Date",
    "Item/Service",
    "Unit Price",
    "Memo",
]
COL_WIDTHS = [16, 28, 38, 12, 32, 14, 30]


def resolve_path(argv_path):
    if argv_path:
        p = Path(argv_path)
        if p.exists():
            return p
        sys.exit("Workbook not found: " + str(p))
    candidates = [
        Path(os.environ.get("USERPROFILE", "")) / "OneDrive" / "Desktop" / "WORK ORDERS" / "RazorSync_Invoice_Tracker.xlsx",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Work Order Tracker" / "RazorSync_Invoice_Tracker.xlsx",
    ]
    for c in candidates:
        if c.exists():
            return c
    sys.exit("Cannot find RazorSync_Invoice_Tracker.xlsx. Pass the path as an argument.")


def rebuild_invoice_import(wb):
    if SHEET_NAME in wb.sheetnames:
        # Preserve sheet position; remove and recreate.
        old_idx = wb.sheetnames.index(SHEET_NAME)
        del wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME, index=old_idx)
    else:
        ws = wb.create_sheet(SHEET_NAME, index=0)

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cell_font = Font(name="Arial", size=10)
    align_left  = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Header row
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(1, i, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Pre-fill body rows: VLOOKUP formula in F, formatting on all 7 cols.
    for r in range(2, ROW_MAX + 1):
        ws.cell(r, 6).value = (
            "=IFERROR(IF(E" + str(r) + "=\"\",\"\","
            "VLOOKUP(E" + str(r) + ",'" + LIB_SHEET + "'!A:C,3,FALSE)),\"\")"
        )
        for c in range(1, 8):
            cell = ws.cell(r, c)
            cell.font = cell_font
            cell.alignment = align_right if c == 6 else align_left
            cell.border = Border(left=thin, right=thin, bottom=thin)
        ws.cell(r, 4).number_format = "MM/DD/YYYY"
        ws.cell(r, 6).number_format = "$#,##0.00;($#,##0.00);-"

    # Data validation: Item/Service dropdown sourced from ServiceItemNames
    # defined name (auto-resizing OFFSET formula maintained on Service Items).
    dv_item = DataValidation(
        type="list",
        formula1="=ServiceItemNames",
        allow_blank=True,
        showDropDown=False,  # show the dropdown arrow
    )
    dv_item.error = "Pick from the Service Items library."
    dv_item.errorTitle = "Invalid item"
    dv_item.add("E2:E" + str(ROW_MAX))
    ws.add_data_validation(dv_item)

    # Customer Name dropdown sourced from the CustomerNames defined name
    # (written into README col D by rebuild_readme). Free text still allowed.
    dv_cust = DataValidation(
        type="list",
        formula1="=" + CUSTOMER_NAMES_RANGE_NAME,
        allow_blank=True,
        showDropDown=False,
    )
    dv_cust.add("B2:B" + str(ROW_MAX))
    ws.add_data_validation(dv_cust)


def rebuild_readme(wb):
    if README_SHEET in wb.sheetnames:
        del wb[README_SHEET]
    ws = wb.create_sheet(README_SHEET, index=len(wb.sheetnames))

    title_font = Font(name="Arial", size=14, bold=True, color="305496")
    h_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    rows = [
        ("RazorSync Invoice Tracker - README", title_font),
        ("", body_font),
        ("Invoice Import sheet", h_font),
        ("Columns A-G:", body_font),
        ("  A  Invoice #         Filled from RazorSync after invoicing.", body_font),
        ("  B  Customer Name     'American Homes 4 Rent', 'Main Street Renewal', or PM verbatim.", body_font),
        ("  C  Service Address   Property service address.", body_font),
        ("  D  Date              Service / invoice date (manual).", body_font),
        ("  E  Item/Service      Pick from Service Items library dropdown.", body_font),
        ("  F  Unit Price        Auto-fills via VLOOKUP from Service Items by item name.", body_font),
        ("  G  Memo              'WO #######' for non-AMH; 'WO ####### | NCXXXX' for AMH.", body_font),
        ("", body_font),
        ("Rows", h_font),
        ("Each WO writes one row per service item plus one blank separator row.", body_font),
        ("Customer Name, Service Address, and Memo are written on the first item row only;", body_font),
        ("subsequent item rows leave those columns blank for readability.", body_font),
        ("The Item/Service dropdown and Unit Price formula extend to row " + str(ROW_MAX) + ".", body_font),
        ("", body_font),
        ("Service Items sheet", h_font),
        ("Columns: Item Name | Description | Price | Taxable (Yes/No).", body_font),
        ("Add new items at the bottom; the ServiceItemNames defined name auto-resizes.", body_font),
        ("All labor is taxed; the Taxable column is informational only.", body_font),
        ("", body_font),
        ("Tracker sync", h_font),
        ("sync_to_lookup.py only writes WOs whose tab is 'Invoiced' in the tracker.", body_font),
        ("Move WOs to the Invoiced tab in the tracker before running Sync Workbook.", body_font),
    ]

    ws.column_dimensions["A"].width = 110
    for i, (text, font) in enumerate(rows, start=1):
        c = ws.cell(i, 1, text)
        c.font = font
        c.alignment = wrap

    # Write customer name list into column D (hidden helper range).
    # The CustomerNames defined name points here so the B-column dropdown
    # sources from a range rather than a hardcoded formula string.
    for i, name in enumerate(CUSTOMER_NAMES, start=1):
        ws.cell(i, 4, name).font = Font(name="Arial", size=10)


def main():
    argv_path = sys.argv[1] if len(sys.argv) > 1 else None
    wb_path = resolve_path(argv_path)
    print("Opening: " + str(wb_path))
    wb = load_workbook(wb_path)

    if LIB_SHEET not in wb.sheetnames:
        sys.exit("Library sheet '" + LIB_SHEET + "' not found - aborting (will not modify workbook).")

    rebuild_invoice_import(wb)
    rebuild_readme(wb)

    # Register (or overwrite) the CustomerNames defined name pointing to the
    # helper list written into README col D by rebuild_readme.
    n_rows = len(CUSTOMER_NAMES)
    ref = "'" + README_SHEET + "'!$" + CUSTOMER_NAMES_COL + "$1:$" + CUSTOMER_NAMES_COL + "$" + str(n_rows)
    if CUSTOMER_NAMES_RANGE_NAME in wb.defined_names:
        del wb.defined_names[CUSTOMER_NAMES_RANGE_NAME]
    wb.defined_names[CUSTOMER_NAMES_RANGE_NAME] = DefinedName(CUSTOMER_NAMES_RANGE_NAME, attr_text=ref)

    tmp_path = wb_path.with_suffix(".tmp.xlsx")
    try:
        wb.save(str(tmp_path))
        _os_replace(str(tmp_path), str(wb_path))
    except Exception as e:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        sys.exit("Failed to save workbook: " + str(e))
    print("Done. 'Invoice Import' and '" + README_SHEET + "' rebuilt at: " + str(wb_path))


if __name__ == "__main__":
    main()
