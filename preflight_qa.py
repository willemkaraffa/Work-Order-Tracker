"""
preflight_qa.py
Dry-run check of the Invoiced-tab work orders before sync. Reports issues
without writing the workbook so you can fix them once instead of every sync.

Issues flagged:
  - No line items resolved (scraper/MSR/AMH all came up empty)
  - Mapping fell back to Labor!/Materials! (no real Service Items match)
  - $0 unit price (Service Items VLOOKUP will return 0)
  - MSR address has no folder match
  - WO id suspicious (>10 digits suggests merged house-number, e.g. 9698891706)
  - Missing customer name or address

Usage:
  python preflight_qa.py                       # auto-detect workbook, text output
  python preflight_qa.py --json                # machine-readable (used by the app's Preflight button)
  python preflight_qa.py "<path-to.xlsx>"     # explicit workbook
"""

import contextlib, io, json, re, sys

from sync_to_lookup import (
    resolve_workbook, resolve_msr_dir, load_orders, load_scraped_items,
    load_overrides, load_service_items, find_msr_folder, extract_msr_items,
    map_to_service_item, customer_name,
)
from openpyxl import load_workbook


WO_ID_MAX_DIGITS = 8  # AMH/MSR WO ids are 7-8 digits; >8 suggests a merged house number (e.g. 9698891706)

FALLBACK_NAMES = {"Labor!", "Materials!"}


def _wo_id_suspicious(wo_id):
    digits = re.sub(r"\D", "", wo_id or "")
    return len(digits) > WO_ID_MAX_DIGITS


def _resolve_raw_items(o, scraped, msr_base):
    """Mirror of write_invoice_import's item-resolution logic. Returns (items, source)."""
    pm      = o.get("pm", "")
    address = o.get("address", "")
    wo_id   = (o.get("id") or "").strip()
    wo_date = o.get("dateCreated", "")

    if wo_id in scraped:
        items = [
            {"name": it.get("name", ""),
             "price": float(it.get("price", 0) or 0),
             "qty": int(it.get("qty", 1) or 1)}
            for it in scraped[wo_id] if it.get("name")
        ]
        return items, "scraped"
    if pm == "MSR" and msr_base:
        folder = find_msr_folder(address, msr_base)
        if folder:
            return extract_msr_items(folder, wo_date), "msr"
        return [], "msr-no-folder"
    if pm == "AMH":
        bid_items = o.get("bidItems") or []
        items = [
            {"name": it.get("name", ""),
             "price": float(it.get("price", 0) or 0),
             "qty": int(it.get("qty", 1) or 1)}
            for it in bid_items if it.get("name")
        ]
        return items, "amh-bid-items" if items else "amh-empty"
    return [], "unknown"


def collect(workbook_path, msr_base):
    """Run all checks and return a dict of issue lists. Pure data, no printing."""
    orders        = load_orders()
    scraped       = load_scraped_items()
    overrides     = load_overrides()
    wb            = load_workbook(workbook_path, read_only=True, data_only=True)
    service_items = load_service_items(wb) if "Service Items" in wb.sheetnames else []
    service_price = {si["name"]: si["price"] for si in service_items}

    no_items      = []
    fallback_only = []
    zero_price    = []
    msr_no_folder = []
    suspicious_id = []
    missing_meta  = []

    for o in orders:
        wo_id   = (o.get("id") or "").strip()
        pm      = o.get("pm", "")
        address = (o.get("address") or "").strip()
        cname   = customer_name(pm)

        if not address or not cname:
            missing_meta.append((wo_id, pm, address, cname))
        if _wo_id_suspicious(wo_id):
            suspicious_id.append((wo_id, address))

        items, source = _resolve_raw_items(o, scraped, msr_base)
        if source == "msr-no-folder":
            msr_no_folder.append((wo_id, address))

        if not items:
            no_items.append((wo_id, pm, address, source))
            continue

        wo_fallback = []
        wo_zero     = []
        for it in items:
            svc = map_to_service_item(it["name"], it["price"], service_items, overrides) if service_items else ""
            if svc in FALLBACK_NAMES:
                wo_fallback.append((it["name"], svc))
            # Zero-price = no override price AND no Service Items match AND raw price is 0.
            unit_price = service_price.get(svc, 0.0) if svc not in FALLBACK_NAMES else it["price"]
            if not unit_price:
                wo_zero.append((it["name"], svc))
        if wo_fallback:
            fallback_only.append((wo_id, address, wo_fallback))
        if wo_zero:
            zero_price.append((wo_id, address, wo_zero))

    return {
        "totalOrders":    len(orders),
        "suspiciousId":   [{"wo": r[0], "address": r[1]} for r in suspicious_id],
        "missingMeta":    [{"wo": r[0], "pm": r[1], "address": r[2], "customer": r[3]} for r in missing_meta],
        "noItems":        [{"wo": r[0], "pm": r[1], "address": r[2], "source": r[3]} for r in no_items],
        "msrNoFolder":    [{"wo": r[0], "address": r[1]} for r in msr_no_folder],
        "fallback":       [{"wo": r[0], "address": r[1], "items": [{"name": n, "mapped": s} for n, s in r[2]]} for r in fallback_only],
        "zeroPrice":      [{"wo": r[0], "address": r[1], "items": [{"name": n, "mapped": s} for n, s in r[2]]} for r in zero_price],
    }


def run_text(workbook_path, msr_base):
    """Human-readable report (terminal)."""
    data = collect(workbook_path, msr_base)

    def section(title, rows, fmt):
        print()
        print("== " + title + " (" + str(len(rows)) + ") ==")
        if not rows:
            print("  (none)")
            return
        for r in rows:
            print("  " + fmt(r))

    section("Suspicious WO ids (>%d digits)" % WO_ID_MAX_DIGITS,
            data["suspiciousId"],
            lambda r: r["wo"] + "  " + (r["address"] or ""))

    section("Missing customer/address",
            data["missingMeta"],
            lambda r: (r["wo"] or "(no id)") + "  pm=" + (r["pm"] or "?") + "  addr='" + (r["address"] or "") + "'  cust='" + (r["customer"] or "") + "'")

    section("No line items resolved",
            data["noItems"],
            lambda r: r["wo"] + "  pm=" + (r["pm"] or "?") + "  src=" + r["source"] + "  " + (r["address"] or ""))

    section("MSR addresses with no folder match",
            data["msrNoFolder"],
            lambda r: r["wo"] + "  " + r["address"])

    section("Mapping fell back to Labor!/Materials!",
            data["fallback"],
            lambda r: r["wo"] + "  " + r["address"] + "\n     " +
                      "\n     ".join(i["name"] + " -> " + i["mapped"] for i in r["items"]))

    section("$0 unit price (will land blank or 0 in workbook)",
            data["zeroPrice"],
            lambda r: r["wo"] + "  " + r["address"] + "\n     " +
                      "\n     ".join(i["name"] + " -> " + (i["mapped"] or "(unmapped)") for i in r["items"]))

    print()
    total_issues = sum(len(data[k]) for k in ("suspiciousId", "missingMeta", "noItems", "msrNoFolder", "fallback", "zeroPrice"))
    print("Preflight complete: " + str(total_issues) + " issue(s) across " + str(data["totalOrders"]) + " WO(s).")


if __name__ == "__main__":
    args      = [a for a in sys.argv[1:] if a]
    want_json = "--json" in args
    args      = [a for a in args if a != "--json"]
    argv_path = args[0] if args else None

    if want_json:
        try:
            # The shared helpers print progress to stdout; capture it so only
            # one valid JSON line reaches the parent process.
            sink = io.StringIO()
            with contextlib.redirect_stdout(sink):
                workbook_path = resolve_workbook(argv_path)
                msr_base      = resolve_msr_dir()
                data          = collect(workbook_path, msr_base)
            print(json.dumps({"ok": True, "data": data, "log": sink.getvalue()}))
        except SystemExit as exc:
            print(json.dumps({"ok": False, "error": str(exc)}))
        except Exception as exc:
            print(json.dumps({"ok": False, "error": type(exc).__name__ + ": " + str(exc)}))
    else:
        workbook_path = resolve_workbook(argv_path)
        msr_base      = resolve_msr_dir()
        run_text(workbook_path, msr_base)
