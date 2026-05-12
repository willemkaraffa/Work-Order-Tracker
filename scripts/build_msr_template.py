"""
Generates docs/MSR_BID_TEMPLATE.xlsx -- the structured bid sheet layout that
sync_to_lookup.py's MSR extractors expect. Hand this to MSR/vendors when asking
for a consistent format. Re-run after layout edits to regenerate.

Usage:  python scripts/build_msr_template.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

OUT = Path(__file__).resolve().parent.parent / "docs" / "MSR_BID_TEMPLATE.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
NOTE_FONT   = Font(italic=True, color="555555")
THIN        = Side(border_style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, cells):
    for addr in cells:
        c = ws[addr]
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = BORDER


def hvac_sheet(wb):
    ws = wb.create_sheet("Vendor HVAC Bid Sheet")
    ws["A1"] = "MSR Vendor HVAC Bid Sheet -- TEMPLATE"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Rows 1-10: free-form header (property, vendor, dates, etc.)"
    ws["A2"].font = NOTE_FONT
    ws["A3"] = "Rows 11-76: ITEMIZED bids. Column B = item description, H = qty, I = line total."
    ws["A3"].font = NOTE_FONT
    ws["A4"] = "Rows 77-81: free-text 'Other' block. Format each line as: $NNN description"
    ws["A4"].font = NOTE_FONT
    ws["A5"] = "Do not rename column headers. Do not insert/delete rows above row 11."
    ws["A5"].font = NOTE_FONT

    # Row 10 = column headers
    ws["A10"] = "#"
    ws["B10"] = "Item / Description"
    ws["C10"] = "Notes (optional)"
    ws["H10"] = "Qty"
    ws["I10"] = "Line Total"
    style_header(ws, ["A10", "B10", "C10", "H10", "I10"])

    # Example rows
    examples = [
        (1, "Replace condensate pump", 1, 185.00),
        (2, "Install new thermostat (Honeywell T6)", 1, 220.00),
        (3, "Labor: drain line clearing", 2, 150.00),
    ]
    for idx, (n, desc, qty, total) in enumerate(examples, start=11):
        ws.cell(idx, 1, n).alignment = CENTER
        ws.cell(idx, 2, desc).alignment = LEFT
        ws.cell(idx, 8, qty).alignment = CENTER
        ws.cell(idx, 9, total).number_format = "$#,##0.00"
        for col in (1, 2, 3, 8, 9):
            ws.cell(idx, col).border = BORDER

    # "Other" prompt
    ws["B77"] = "Please provide any other items below in the format: $NNN description"
    ws["B77"].font = NOTE_FONT
    ws["C77"] = "$95 freon top-off\n$60 disposal fee"
    ws["C77"].alignment = LEFT

    # Bid total marker (parser skips rows starting with TOTAL/BID TOTAL)
    ws["B83"] = "BID TOTAL"
    ws["B83"].font = Font(bold=True)
    ws["I83"] = "=SUM(I11:I81)"
    ws["I83"].number_format = "$#,##0.00"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 14


def plumbing_sheet(wb):
    ws = wb.create_sheet("Plumbing")
    ws["A1"] = "MSR Plumbing Bid Sheet -- TEMPLATE"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Rows 1-10: free-form header."
    ws["A2"].font = NOTE_FONT
    ws["A3"] = "Rows 11+: column C = item name, D = free-text description (Other rows only), I = qty, J = line total."
    ws["A3"].font = NOTE_FONT
    ws["A4"] = "For 'Other' rows, put the literal text 'Other (please provide)' in C and the $NNN description list in D."
    ws["A4"].font = NOTE_FONT

    ws["A10"] = "#"
    ws["C10"] = "Item / Description"
    ws["D10"] = "Notes / Other detail"
    ws["I10"] = "Qty"
    ws["J10"] = "Line Total"
    style_header(ws, ["A10", "C10", "D10", "I10", "J10"])

    examples = [
        (1, "Replace water heater (40 gal gas)", "", 1, 1450.00),
        (2, "Labor: snake main drain", "", 1, 275.00),
        (3, "Other (please provide)", "$120 wax ring replacement\n$45 supply lines x2", 1, 165.00),
    ]
    for idx, (n, name, detail, qty, total) in enumerate(examples, start=11):
        ws.cell(idx, 1, n).alignment = CENTER
        ws.cell(idx, 3, name).alignment = LEFT
        ws.cell(idx, 4, detail).alignment = LEFT
        ws.cell(idx, 9, qty).alignment = CENTER
        ws.cell(idx, 10, total).number_format = "$#,##0.00"
        for col in (1, 3, 4, 9, 10):
            ws.cell(idx, col).border = BORDER

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["I"].width = 8
    ws.column_dimensions["J"].width = 14


def readme_sheet(wb):
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "MSR Bid Sheet Template"
    ws["A1"].font = Font(bold=True, size=16)
    notes = [
        "",
        "Purpose: a consistent file layout vendors can fill out so the Work Order Tracker can auto-import line items.",
        "",
        "Sheets:",
        "  - Vendor HVAC Bid Sheet : standard HVAC bid format",
        "  - Plumbing              : plumbing-specific layout",
        "",
        "Rules that must not change:",
        "  1. The sheet names ('Vendor HVAC Bid Sheet' / 'Plumbing') -- the parser dispatches by name.",
        "  2. Column positions of qty, line total, item description, and the 'Other' free-text block.",
        "  3. Row 11 is the first item row on both sheets.",
        "  4. CO files (change orders) should be named with 'CO' or 'Change Order' / 'Change Request' in the filename.",
        "  5. Save each property under: WORK ORDERS\\aMain Street Renewal\\<HouseNumber Street Name>\\<file>.xlsx",
        "",
        "What gets imported:",
        "  - Description (column B on HVAC, C on Plumbing)",
        "  - Qty",
        "  - Unit price (= line total / qty)",
        "  - Free-text 'Other' lines of the form '$NNN description'",
    ]
    for i, line in enumerate(notes, start=2):
        ws.cell(i, 1, line).alignment = LEFT
    ws.column_dimensions["A"].width = 110


wb = Workbook()
del wb[wb.sheetnames[0]]  # drop default
readme_sheet(wb)
hvac_sheet(wb)
plumbing_sheet(wb)
wb.save(OUT)
print("Wrote " + str(OUT))
