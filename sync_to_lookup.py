"""
sync_to_lookup.py
Reads work orders from wo-data.json and writes Invoiced-tab orders into the
'Invoice Import' sheet of RazorSync_Invoice_Tracker.xlsx.

Auto-population pipeline:
  Pre-scraped  — reads scraped_bid_items.json if present (written by the
                 /sync-wo-tracker Cowork skill); takes priority for any WO
                 whose id appears in the file.
  MSR WOs      — locates matching job spreadsheet under WORK ORDERS\\aMain Street Renewal\\,
                 filters to files within 30 days of the WO date, handles CO
                 sheets, deduplicates items across sheets.
  AMH WOs      — reads bidItems captured by the Chrome extension at import time;
                 if absent, a blank anchor row is written (run the Cowork skill
                 first to populate scraped_bid_items.json).
  Fallback     — single blank anchor row (user fills Item/Service manually).

Layout (RazorSync-import-ready, 7 columns):
    A: Invoice #     - blank (filled from RazorSync after invoice creation)
    B: Customer Name - 'American Homes 4 Rent' / 'Main Street Renewal' / pm verbatim
    C: Service Address
    D: Date          - blank (manual)
    E: Item/Service  - mapped service item name (or blank if no match)
    F: Unit Price    - VLOOKUP formula returning Price from Service Items (col C)
    G: Memo          - 'WO #######' / 'WO ####### | NCXXXX' for AMH

Multi-item WOs: B, C, G are written only on the first item row; subsequent
item rows leave those columns blank for readability.

Requirements: pip install openpyxl
"""

import json, os, re, sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

HERE         = Path(__file__).parent
WO_JSON      = Path(os.environ.get("APPDATA", "")) / "work-order-tracker" / "wo-data.json"
SCRAPED_JSON = Path(os.environ.get("APPDATA", "")) / "work-order-tracker" / "scraped_bid_items.json"

SHEET_NAME = "Invoice Import"

CUSTOMER_MAP = {
    "AMH": "American Homes 4 Rent",
    "MSR": "Main Street Renewal",
}

WRITE_ROW_MAX = 2000

CELL_FONT   = Font(name="Arial", size=10)
ALIGN_LEFT  = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
THIN        = Side(border_style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, bottom=THIN)

# Street type normalisation (expand or abbreviate -> canonical abbreviation)
_STREET_TYPES = {
    'lane':'ln','drive':'dr','court':'ct','street':'st','avenue':'ave',
    'boulevard':'blvd','circle':'cir','place':'pl','road':'rd','way':'way',
    'trail':'trl','terrace':'ter','parkway':'pkwy','highway':'hwy',
    'ln':'ln','dr':'dr','ct':'ct','st':'st','ave':'ave','blvd':'blvd',
    'cir':'cir','pl':'pl','rd':'rd','trl':'trl','ter':'ter','pkwy':'pkwy',
}

_LABOR_KW    = {'labor','install','replace','repair','clean','clear','service',
                'check','diagnos','flush','seal','secure','rebuild','auger',
                'pump','dig','remove','inspect'}
_MATERIAL_KW = {'material','mat','mat-','part','unit','equipment','supply',
                'supplies','filter','capacitor','contactor','valve','coil',
                'motor','thermostat','fuse','sensor','pump','switch','board'}

_STOP_WORDS  = {'','to','the','a','an','of','in','and','or','with','for','is',
                'are','at','by','on','it','its','this','that','per','new',
                'existing','old','current'}

# CO file detection: "CO" as a whole word, or "change request"/"change order" substring
_CO_RE = re.compile(r'\bco\b|change\s+request|change\s+order', re.IGNORECASE)


# ── Tokeniser (shared by mapping and dedup) ───────────────────────────────────

def _tokenise(text):
    return frozenset(t for t in re.split(r"\W+", text.lower()) if t and t not in _STOP_WORDS)


# ── Workbook resolution ────────────────────────────────────────────────────────

def resolve_workbook(argv_path):
    if argv_path:
        p = Path(argv_path)
        if p.exists():
            return p
        sys.exit("Workbook not found at provided path:\n  " + str(p))
    candidates = [
        Path(os.environ.get("USERPROFILE", "")) / "OneDrive" / "Desktop" / "WORK ORDERS" / "RazorSync_Invoice_Tracker.xlsx",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Work Order Tracker" / "RazorSync_Invoice_Tracker.xlsx",
        HERE / "RazorSync_Invoice_Tracker.xlsx",
    ]
    for c in candidates:
        if c.exists():
            return c
    sys.exit(
        "Cannot find RazorSync_Invoice_Tracker.xlsx.\n"
        "Run: python sync_to_lookup.py \"C:\\path\\to\\RazorSync_Invoice_Tracker.xlsx\""
    )


def resolve_msr_dir():
    candidates = [
        Path(os.environ.get("USERPROFILE", "")) / "OneDrive" / "Desktop" / "WORK ORDERS" / "aMain Street Renewal",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Work Order Tracker" / "aMain Street Renewal",
        HERE / "aMain Street Renewal",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


# ── WO data ───────────────────────────────────────────────────────────────────

def load_orders():
    if not WO_JSON.exists():
        sys.exit("Cannot find wo-data.json at:\n  " + str(WO_JSON))
    raw    = json.loads(WO_JSON.read_text(encoding="utf-8"))
    data   = json.loads(raw["wo_data"])
    all_orders = data.get("orders", [])
    invoiced = [
        o for o in all_orders
        if not o.get("deleted", False) and o.get("tab") == "invoiced"
    ]
    print("Loaded " + str(len(all_orders)) + " orders total; " + str(len(invoiced)) + " in Invoiced tab")
    return invoiced


# ── Pre-scraped items (written by Cowork skill) ───────────────────────────────

def load_scraped_items():
    if not SCRAPED_JSON.exists():
        return {}
    try:
        data = json.loads(SCRAPED_JSON.read_text(encoding="utf-8"))
        items = data.get("items", {})
        print("Loaded pre-scraped items for " + str(len(items)) + " WO(s) from scraped_bid_items.json")
        return items
    except Exception as e:
        print("[WARN] Could not read scraped_bid_items.json: " + str(e))
        return {}


# ── Service Items library ─────────────────────────────────────────────────────

def load_service_items(wb):
    if "Service Items" not in wb.sheetnames:
        return []
    ws = wb["Service Items"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0] if row else None
        desc = row[1] if len(row) > 1 else None
        price = row[2] if len(row) > 2 else None
        if name and isinstance(name, str) and name.strip():
            items.append({
                "name":  name.strip(),
                "desc":  (desc or "").strip(),
                "price": float(price) if isinstance(price, (int, float)) else 0.0,
            })
    return items


# ── Address / folder matching for MSR ────────────────────────────────────────

def _addr_tokens(text):
    """Return (house_number_str, frozenset_of_street_tokens) from an address or folder name."""
    street = text.split(",")[0].strip().lower()
    parts  = re.split(r"[\s\-/]+", street)
    num    = None
    tokens = []
    for p in parts:
        if not p:
            continue
        if p.isdigit() and num is None:
            num = p
        else:
            tokens.append(_STREET_TYPES.get(p, p))
    return num, frozenset(t for t in tokens if len(t) > 1)


def find_msr_folder(address, msr_base):
    if not msr_base or not Path(msr_base).exists():
        return None
    wo_num, wo_tok = _addr_tokens(address)
    if not wo_num:
        return None

    best_score  = 0
    best_folder = None
    for folder in Path(msr_base).iterdir():
        if not folder.is_dir():
            continue
        f_num, f_tok = _addr_tokens(folder.name)
        if f_num != wo_num:
            continue
        overlap = len(wo_tok & f_tok)
        if overlap > best_score:
            best_score  = overlap
            best_folder = folder

    return best_folder if best_score > 0 else None


# ── MSR spreadsheet item extraction ──────────────────────────────────────────

_SKIP_B = {"TOTAL ", "BID TOTAL", "ITEM"}

def _skip_row(b_val):
    if not b_val:
        return True
    s = str(b_val).strip().upper()
    return any(s.startswith(p) for p in _SKIP_B) or s == "ITEM"


def _parse_other_block(text):
    """Parse 'Other' free-text block: lines/segments of the form '$NNN description'."""
    items = []
    for segment in re.split(r"\n", text):
        segment = segment.strip()
        if not segment:
            continue
        parts = re.split(r"(?=\$\d)", segment)
        for part in parts:
            part = part.strip()
            m = re.match(r"\$([\d,]+(?:\.\d+)?)\s+(.+)", part)
            if m:
                price = float(m.group(1).replace(",", ""))
                name  = m.group(2).strip()
                name = re.split(r"\s*\$\d", name)[0].strip()
                if name:
                    items.append({"name": name, "price": price, "qty": 1})
    return items


def _is_co_file(stem):
    return bool(_CO_RE.search(stem))


def _file_mtime(path):
    return datetime.fromtimestamp(path.stat().st_mtime)


def _items_duplicate(a, b):
    """True when two items represent the same charge: similar name AND price within 10%."""
    tok_a = _tokenise(a["name"])
    tok_b = _tokenise(b["name"])
    if not tok_a or not tok_b:
        return False
    union = len(tok_a | tok_b)
    if union == 0 or len(tok_a & tok_b) / union < 0.5:
        return False
    pa, pb = a["price"], b["price"]
    if pa > 0 and pb > 0:
        return abs(pa - pb) / max(pa, pb) <= 0.10
    return True


def _extract_hvac_sheet(ws):
    """Extract items from a 'Vendor HVAC Bid Sheet' worksheet."""
    items = []
    for row in ws.iter_rows(min_row=11, max_row=76, values_only=True):
        if len(row) < 9:
            continue
        b_val, h_val, i_val = row[1], row[7], row[8]
        if _skip_row(b_val):
            continue
        if not isinstance(i_val, (int, float)) or i_val <= 0:
            continue
        qty = int(h_val) if isinstance(h_val, (int, float)) and h_val > 0 else 1
        unit_price = round(float(i_val) / qty, 2) if qty else float(i_val)
        items.append({"name": str(b_val).strip(), "price": unit_price, "qty": qty})
    # Other free-text block (rows 77-81): $NNN description in col C
    for row in ws.iter_rows(min_row=77, max_row=81, values_only=True):
        if len(row) < 9:
            continue
        b_val, c_val, i_val = row[1], row[2], row[8]
        if (c_val and isinstance(b_val, str)
                and "please provide" in str(b_val).lower()
                and isinstance(i_val, (int, float)) and i_val > 0):
            items.extend(_parse_other_block(str(c_val)))
    return items


def _extract_plumbing_sheet(ws):
    """Extract items from a 'Plumbing' worksheet.

    Column layout: C=item name (idx 2), D=description/free-text (idx 3),
    I=qty (idx 8), J=line item price (idx 9).
    """
    items = []
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        if len(row) < 10:
            continue
        c_val = row[2]   # item name
        d_val = row[3]   # free-text description (Other rows)
        j_val = row[9]   # line item price
        if _skip_row(c_val):
            continue
        if not isinstance(j_val, (int, float)) or j_val <= 0:
            continue
        name = str(c_val).strip()
        # "Other" row: parse the description column as $NNN free-text
        if "other" in name.lower() and "please provide" in name.lower() and d_val:
            items.extend(_parse_other_block(str(d_val)))
        else:
            items.append({"name": name, "price": float(j_val), "qty": 1})
    return items


def _extract_from_file(xlsx_path):
    try:
        wb = load_workbook(str(xlsx_path), data_only=True)
    except Exception as e:
        print("  [MSR] Could not open " + xlsx_path.name + ": " + str(e))
        return []
    ws = wb.active
    if "plumbing" in (ws.title or "").lower():
        return _extract_plumbing_sheet(ws)
    return _extract_hvac_sheet(ws)


def extract_msr_items(folder_path, wo_date=None):
    folder = Path(folder_path)
    xlsx_files = [f for f in folder.glob("*.xlsx") if "~$" not in f.name]
    if not xlsx_files:
        print("  [MSR] No xlsx found in: " + str(folder))
        return []

    cutoff = None
    if wo_date:
        try:
            cutoff = datetime.strptime(wo_date, "%Y-%m-%d")
        except Exception:
            pass

    def _within_window(f):
        if cutoff is None:
            return True
        return abs((_file_mtime(f) - cutoff).days) <= 30

    originals = [f for f in xlsx_files if not _is_co_file(f.stem)]
    cos       = [f for f in xlsx_files if _is_co_file(f.stem)]

    in_window_orig = [f for f in originals if _within_window(f)]
    in_window_cos  = [f for f in cos if _within_window(f)]

    # If date filter eliminated everything, fall back to all files
    if not in_window_orig and not in_window_cos:
        print("  [MSR] No files within 30 days of WO date -- using all files")
        in_window_orig = originals
        in_window_cos  = cos

    in_window_cos.sort(key=lambda f: f.stat().st_mtime)

    all_items = []

    def _add_deduped(new_items, label):
        added = 0
        for item in new_items:
            if any(_items_duplicate(item, existing) for existing in all_items):
                continue
            all_items.append(item)
            added += 1
        print("  [MSR] " + label + " -> " + str(added) + " item(s) added")

    for f in in_window_orig:
        _add_deduped(_extract_from_file(f), f.name)

    for f in in_window_cos:
        _add_deduped(_extract_from_file(f), "[CO] " + f.name)

    print("  [MSR] Total after dedup: " + str(len(all_items)))
    return all_items


# ── Service item mapping ──────────────────────────────────────────────────────

def map_to_service_item(name, price, service_items):
    """
    Return the best-matching service item name from the library, or 'Labor!' / 'Materials!'.
    Matching: exact -> token Jaccard (boosted when price is close) -> keyword fallback.
    """
    name_lower = name.lower()
    name_tok   = _tokenise(name)

    best_score = 0.0
    best_name  = None

    for si in service_items:
        si_name = si["name"]
        if si_name in ("Labor!", "Materials!"):
            continue
        si_lower = si_name.lower()
        if name_lower == si_lower:
            return si_name
        si_tok = _tokenise(si_name) | _tokenise(si["desc"])
        if not name_tok or not si_tok:
            continue
        overlap = len(name_tok & si_tok)
        union   = len(name_tok | si_tok)
        jaccard = overlap / union if union else 0.0
        price_boost = 0.0
        if si["price"] and price:
            if abs(si["price"] - price) / max(abs(price), 1) < 0.15:
                price_boost = 0.25
        score = jaccard + price_boost
        if score > best_score:
            best_score = score
            best_name  = si_name

    if best_score >= 0.25 and best_name:
        return best_name

    toks = name_tok
    if toks & _MATERIAL_KW:
        return "Materials!"
    if toks & _LABOR_KW:
        return "Labor!"
    return "Labor!"


# ── Formatting helpers ────────────────────────────────────────────────────────

def customer_name(pm):
    pm = (pm or "").strip()
    return CUSTOMER_MAP.get(pm, pm)


def memo_for(o):
    wo = (o.get("id") or "").strip()
    if wo:
        base = wo if wo.upper().startswith("WO") else "WO " + wo
    else:
        base = ""
    if o.get("pm") == "AMH":
        pid = (o.get("propertyId") or "").strip()
        if pid:
            return (base + " | " + pid) if base else pid
    return base


def style_row(ws, row):
    for c in range(1, 8):
        cell = ws.cell(row, c)
        cell.font      = CELL_FONT
        cell.border    = CELL_BORDER
        cell.alignment = ALIGN_RIGHT if c == 6 else ALIGN_LEFT
    ws.cell(row, 4).number_format = "MM/DD/YYYY"
    ws.cell(row, 6).number_format = "$#,##0.00;($#,##0.00);-"


def fill_unit_price_formula(ws, row):
    ws.cell(row, 6).value = (
        "=IFERROR(IF(E" + str(row) + "=\"\",\"\","
        "VLOOKUP(E" + str(row) + ",'Service Items'!A:C,3,FALSE)),\"\")"
    )


# ── Main write ────────────────────────────────────────────────────────────────

def write_invoice_import(orders, workbook_path, msr_base=None):
    scraped = load_scraped_items()

    wb = load_workbook(workbook_path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit("Sheet '" + SHEET_NAME + "' not found in " + workbook_path.name)
    ws = wb[SHEET_NAME]

    service_items = load_service_items(wb)
    if not service_items:
        print("[WARN] Service Items sheet empty or missing -- item mapping disabled.")

    rows_needed  = max(len(orders) * 12, 2)
    clear_up_to  = min(max(ws.max_row, rows_needed, 2), WRITE_ROW_MAX)

    for r in range(2, clear_up_to + 1):
        for c in range(1, 8):
            ws.cell(r, c).value = None

    row = 2

    for o in orders:
        pm      = o.get("pm", "")
        address = o.get("address", "")
        wo_id   = (o.get("id") or "").strip()
        wo_date = o.get("dateCreated", "")
        cname   = customer_name(pm)
        memo    = memo_for(o)

        # --- Resolve line items ---
        raw_items = []

        if wo_id in scraped:
            raw_items = [
                {
                    "name":  it.get("name", ""),
                    "price": float(it.get("price", 0) or 0),
                    "qty":   int(it.get("qty", 1) or 1),
                }
                for it in scraped[wo_id]
                if it.get("name")
            ]
            print("  [Scraped] " + str(len(raw_items)) + " item(s) for " + wo_id)
        elif pm == "MSR" and msr_base:
            folder = find_msr_folder(address, msr_base)
            if folder:
                print("  [MSR] Matched '" + address + "' -> " + folder.name)
                raw_items = extract_msr_items(folder, wo_date)
            else:
                print("  [MSR] No folder match for: " + address)
        elif pm == "AMH":
            bid_items = o.get("bidItems") or []
            if bid_items:
                raw_items = [
                    {
                        "name":  item.get("name", ""),
                        "price": float(item.get("price", 0) or 0),
                        "qty":   int(item.get("qty", 1) or 1),
                    }
                    for item in bid_items
                    if item.get("name")
                ]
                print("  [AMH] " + str(len(raw_items)) + " bid item(s) from extension capture")
            else:
                print("  [AMH] No bidItems for " + wo_id + " -- run /sync-wo-tracker skill first")

        # --- Write rows ---
        if raw_items:
            for i, item in enumerate(raw_items):
                first    = (i == 0)
                svc_name = map_to_service_item(item["name"], item["price"], service_items) if service_items else ""
                ws.cell(row, 1).value = ""
                ws.cell(row, 2).value = cname   if first else ""
                ws.cell(row, 3).value = address if first else ""
                ws.cell(row, 4).value = ""
                ws.cell(row, 5).value = svc_name
                fill_unit_price_formula(ws, row)
                ws.cell(row, 7).value = memo if first else ""
                style_row(ws, row)
                row += 1
            # Blank separator between WOs
            fill_unit_price_formula(ws, row)
            style_row(ws, row)
            row += 1
        else:
            # Anchor row: user fills Item/Service manually
            ws.cell(row, 1).value = ""
            ws.cell(row, 2).value = cname
            ws.cell(row, 3).value = address
            ws.cell(row, 4).value = ""
            ws.cell(row, 5).value = ""
            fill_unit_price_formula(ws, row)
            ws.cell(row, 7).value = memo
            style_row(ws, row)
            row += 1
            fill_unit_price_formula(ws, row)
            style_row(ws, row)
            row += 1

        if row >= WRITE_ROW_MAX - 2:
            print("[WARN] Approaching WRITE_ROW_MAX -- some orders may be truncated.")
            break

    # Restore VLOOKUP across remaining cleared rows
    for r in range(row, clear_up_to + 1):
        fill_unit_price_formula(ws, r)
        style_row(ws, r)

    # Delete scraped file after successful use to avoid stale data on next sync
    if scraped and SCRAPED_JSON.exists():
        try:
            SCRAPED_JSON.unlink()
            print("Deleted scraped_bid_items.json")
        except Exception:
            pass

    # Atomic save
    tmp_path = workbook_path.with_suffix(".tmp.xlsx")
    try:
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(workbook_path))
    except Exception as e:
        try:
            tmp_path.unlink()
        except Exception:
            pass
        sys.exit("Failed to save workbook: " + str(e))

    print("Done -- " + str(len(orders)) + " WO group(s) written to '" + SHEET_NAME + "' in " + workbook_path.name)


if __name__ == "__main__":
    argv_path = sys.argv[1] if len(sys.argv) > 1 else None
    workbook_path = resolve_workbook(argv_path)
    msr_base      = resolve_msr_dir()
    if msr_base:
        print("MSR folder: " + str(msr_base))
    else:
        print("[WARN] MSR folder not found -- MSR item extraction disabled.")
    orders = load_orders()
    write_invoice_import(orders, workbook_path, msr_base)
