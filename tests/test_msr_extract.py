"""
Unit tests for MSR bid sheet extraction.

Run from the project root:
    python -m unittest tests/test_msr_extract.py -v
"""
import sys, tempfile, unittest
from pathlib import Path
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from sync_to_lookup import (
    _extract_hvac_sheet, _extract_plumbing_sheet,
    _parse_other_block, _is_co_file, extract_msr_items,
)


def make_hvac_ws():
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor HVAC Bid Sheet"
    # Row 11+: B=item, H=qty, I=line total
    rows = [
        (11, "B", "Replace condensate pump"), (11, "H", 1), (11, "I", 185.0),
        (12, "B", "Install thermostat"),       (12, "H", 1), (12, "I", 220.0),
        (13, "B", "Labor: drain clear"),       (13, "H", 2), (13, "I", 150.0),  # unit = 75
        (14, "B", "BID TOTAL"),                (14, "I", 555.0),                  # skipped
    ]
    for r, col, val in rows:
        ws[col + str(r)] = val
    # Other block: row 77 col B is prompt; col C contains $NNN free-text
    ws["B77"] = "Please provide other items below"
    ws["C77"] = "$95 freon top-off\n$60 disposal fee"
    return ws


def make_plumbing_ws():
    wb = Workbook()
    ws = wb.active
    ws.title = "Plumbing"
    # C=name, D=detail, I=qty, J=line total
    rows = [
        (11, "C", "Water heater install"), (11, "I", 1), (11, "J", 1450.0),
        (12, "C", "Labor: snake drain"),   (12, "I", 1), (12, "J", 275.0),
        (13, "C", "Other (please provide)"), (13, "I", 1), (13, "J", 165.0),
    ]
    for r, col, val in rows:
        ws[col + str(r)] = val
    ws["D13"] = "$120 wax ring\n$45 supply lines"
    return ws


class HVACExtractTests(unittest.TestCase):

    def test_basic_itemized_rows(self):
        ws    = make_hvac_ws()
        items = _extract_hvac_sheet(ws)
        names = [i["name"] for i in items]
        self.assertIn("Replace condensate pump", names)
        self.assertIn("Install thermostat", names)
        self.assertIn("Labor: drain clear", names)

    def test_unit_price_from_line_total_divided_by_qty(self):
        ws    = make_hvac_ws()
        items = _extract_hvac_sheet(ws)
        drain = next(i for i in items if i["name"] == "Labor: drain clear")
        self.assertEqual(drain["qty"], 2)
        self.assertEqual(drain["price"], 75.0)

    def test_bid_total_row_skipped(self):
        ws    = make_hvac_ws()
        items = _extract_hvac_sheet(ws)
        self.assertFalse(any("BID TOTAL" in i["name"].upper() for i in items))

    def test_other_block_parsed(self):
        ws    = make_hvac_ws()
        items = _extract_hvac_sheet(ws)
        names = [i["name"] for i in items]
        self.assertIn("freon top-off", names)
        self.assertIn("disposal fee", names)


class PlumbingExtractTests(unittest.TestCase):

    def test_basic_plumbing_rows(self):
        ws    = make_plumbing_ws()
        items = _extract_plumbing_sheet(ws)
        names = [i["name"] for i in items]
        self.assertIn("Water heater install", names)
        self.assertIn("Labor: snake drain", names)

    def test_other_row_parses_description(self):
        ws    = make_plumbing_ws()
        items = _extract_plumbing_sheet(ws)
        names = [i["name"] for i in items]
        self.assertIn("wax ring", names)
        self.assertIn("supply lines", names)


class OtherBlockParserTests(unittest.TestCase):

    def test_dollar_prefixed_items_extracted(self):
        items = _parse_other_block("$95 freon top-off\n$60 disposal fee")
        self.assertEqual(len(items), 2)
        self.assertEqual(items[0]["price"], 95.0)
        self.assertEqual(items[1]["price"], 60.0)

    def test_thousand_separator_handled(self):
        items = _parse_other_block("$1,250 ductwork replacement")
        self.assertEqual(items[0]["price"], 1250.0)

    def test_multiple_items_on_one_line(self):
        items = _parse_other_block("$25 part-A $40 part-B")
        names = [i["name"] for i in items]
        self.assertIn("part-A", names)
        self.assertIn("part-B", names)


class COFileDetectionTests(unittest.TestCase):

    def test_co_word_match(self):
        self.assertTrue(_is_co_file("123 Main St CO bid"))
        self.assertTrue(_is_co_file("change request April"))
        self.assertTrue(_is_co_file("Change Order final"))

    def test_non_co(self):
        self.assertFalse(_is_co_file("HVAC bid sheet"))
        self.assertFalse(_is_co_file("123 Main St original"))


class FolderExtractionDedupTests(unittest.TestCase):

    def test_duplicate_items_across_files_deduplicated(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            # File 1: original bid
            wb1 = Workbook()
            ws1 = wb1.active
            ws1.title = "Vendor HVAC Bid Sheet"
            ws1["B11"] = "Replace condensate pump"
            ws1["H11"] = 1
            ws1["I11"] = 185.0
            wb1.save(folder / "bid_original.xlsx")
            # File 2: CO with same item at same price -- should be deduped
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = "Vendor HVAC Bid Sheet"
            ws2["B11"] = "Replace condensate pump"
            ws2["H11"] = 1
            ws2["I11"] = 185.0
            wb2.save(folder / "CO bid_changes.xlsx")

            items = extract_msr_items(folder, wo_date=None)
            names = [i["name"] for i in items]
            self.assertEqual(names.count("Replace condensate pump"), 1)

    def test_distinct_items_preserved(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            wb1 = Workbook()
            ws1 = wb1.active
            ws1.title = "Vendor HVAC Bid Sheet"
            ws1["B11"] = "Replace condensate pump"; ws1["H11"] = 1; ws1["I11"] = 185.0
            ws1["B12"] = "Install thermostat";       ws1["H12"] = 1; ws1["I12"] = 220.0
            wb1.save(folder / "bid.xlsx")

            items = extract_msr_items(folder, wo_date=None)
            names = [i["name"] for i in items]
            self.assertIn("Replace condensate pump", names)
            self.assertIn("Install thermostat", names)
            self.assertEqual(len(items), 2)


if __name__ == "__main__":
    unittest.main()
